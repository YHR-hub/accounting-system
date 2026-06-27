"""FastAPI 后端（阶段3）。

复用 accsys 业务核心（纯计算 + 报表 *_data）与 SQLAlchemy ORM 模型，
对外提供现代 JSON API，并自动生成 OpenAPI/Swagger 文档（/docs）。

运行：uvicorn api.main:app --reload
文档：http://localhost:8000/docs
"""
from __future__ import annotations

import io
from contextlib import asynccontextmanager
from datetime import datetime
from typing import Dict, List

from fastapi import Depends, FastAPI, HTTPException, Query, Response
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import func, select
from sqlalchemy.orm import Session

import accsys as acc
from accsys import repo
from accsys.db import SessionLocal, engine, get_database_url
from accsys.models import JournalEntry, Voucher

from .auth import create_access_token, get_current_user, require_roles, verify_user
from .schemas import (
    AccountCreate,
    AccountOut,
    AiQueryRequest,
    AiQueryResponse,
    EmployeeCreate,
    EmployeeUpdate,
    FixedAssetCreate,
    Health,
    InventoryMove,
    LoginRequest,
    PayrollRun,
    PitResult,
    ProductCreate,
    ProjectCreate,
    ReportRow,
    TokenResponse,
    UserOut,
    VatResult,
    VoucherCreate,
    VoucherCreated,
    VoucherOut,
)


@asynccontextmanager
async def lifespan(app: FastAPI):
    repo.bootstrap(engine)
    yield


app = FastAPI(title="会计系统 API", version="0.1.0", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:8080"],
    allow_methods=["*"],
    allow_headers=["*"],
)


def get_db() -> Session:
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


@app.get("/api/health", response_model=Health, tags=["系统"])
def health():
    url = get_database_url()
    backend = "postgresql" if url.startswith("postgres") else "sqlite"
    return Health(status="ok", database=backend)


@app.get("/api/ratios", tags=["报表"])
def ratios(db: Session = Depends(get_db)) -> Dict[str, float]:
    return repo.financial_ratios(db)


@app.get("/api/accounts", response_model=List[AccountOut], tags=["科目"])
def accounts(db: Session = Depends(get_db)):
    accts = repo.load_accounts(db)
    balances = repo.calc_balances(db, accts)
    rows = db.execute(
        select(
            JournalEntry.account_code,
            func.coalesce(func.sum(JournalEntry.debit), 0),
            func.coalesce(func.sum(JournalEntry.credit), 0),
        ).group_by(JournalEntry.account_code)
    ).all()
    mv = {r[0]: (float(r[1]), float(r[2])) for r in rows}
    out = []
    for a in accts:
        dt, ct = mv.get(a["code"], (0.0, 0.0))
        out.append(AccountOut(
            code=a["code"], name=a["name"], category=a["category"],
            debit=dt, credit=ct, balance=float(balances.get(a["code"], 0)),
        ))
    return out


@app.get("/api/vouchers", response_model=List[VoucherOut], tags=["凭证"])
def vouchers(
    year: int = Query(2026),
    month: int = Query(0, ge=0, le=12),
    db: Session = Depends(get_db),
):
    totals = dict(db.execute(
        select(JournalEntry.voucher_id, func.coalesce(func.sum(JournalEntry.debit), 0))
        .group_by(JournalEntry.voucher_id)
    ).all())

    stmt = select(Voucher).where(Voucher.fiscal_year == year)
    if month > 0:
        stmt = stmt.where(Voucher.fiscal_month == month)
    stmt = stmt.order_by(Voucher.id.desc())

    out = []
    for v in db.execute(stmt).scalars().all():
        out.append(VoucherOut(
            id=v.id, voucher_no=v.voucher_no, date=v.date, summary=v.summary,
            fiscal_year=v.fiscal_year, fiscal_month=v.fiscal_month,
            total=float(totals.get(v.id, 0) or 0),
        ))
    return out


@app.get("/api/reports/balance", response_model=List[ReportRow], tags=["报表"])
def report_balance(db: Session = Depends(get_db)):
    return repo.balance_sheet_data(db)


@app.get("/api/reports/income", response_model=List[ReportRow], tags=["报表"])
def report_income(db: Session = Depends(get_db)):
    return repo.income_statement_data(db)


@app.get("/api/reports/cashflow", response_model=List[ReportRow], tags=["报表"])
def report_cashflow(db: Session = Depends(get_db)):
    return repo.cash_flow_statement_data(db)


@app.get("/api/tax/vat", response_model=VatResult, tags=["税务"])
def tax_vat(revenue: float = Query(..., gt=0), rate: float = Query(0.13)):
    return acc.compute_vat(revenue, rate)


@app.get("/api/tax/pit", response_model=PitResult, tags=["税务"])
def tax_pit(income: float = Query(..., description="月应纳税所得额")):
    return acc.compute_pit(income)


# ── 鉴权 ────────────────────────────────────────────
@app.post("/api/auth/login", response_model=TokenResponse, tags=["鉴权"])
def login(payload: LoginRequest):
    user = verify_user(payload.username, payload.password)
    if not user:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    return TokenResponse(access_token=create_access_token(user), user=UserOut(**user))


@app.get("/api/auth/me", response_model=UserOut, tags=["鉴权"])
def me(user: dict = Depends(get_current_user)):
    return UserOut(**user)


# ── 凭证写入（需 accountant / admin 角色） ─────────────
@app.post("/api/vouchers", response_model=VoucherCreated, status_code=201, tags=["凭证"])
def create_voucher(
    payload: VoucherCreate,
    db: Session = Depends(get_db),
    user: dict = Depends(require_roles("admin", "accountant")),
):
    if len(payload.entries) < 2:
        raise HTTPException(status_code=400, detail="凭证至少需要两条分录")
    for e in payload.entries:
        if e.debit < 0 or e.credit < 0:
            raise HTTPException(status_code=400, detail="借贷金额不能为负")
    total_debit = round(sum(e.debit for e in payload.entries), 2)
    total_credit = round(sum(e.credit for e in payload.entries), 2)
    if total_debit <= 0:
        raise HTTPException(status_code=400, detail="凭证金额必须大于 0")
    if abs(total_debit - total_credit) > 0.01:
        raise HTTPException(status_code=400, detail=f"借贷不平：借方 {total_debit} ≠ 贷方 {total_credit}")
    try:
        dt = datetime.strptime(payload.date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="日期格式应为 YYYY-MM-DD")

    v_no = repo.next_voucher_no(db, dt.year, dt.month)
    voucher = Voucher(
        voucher_no=v_no, date=payload.date, summary=payload.summary,
        fiscal_year=dt.year, fiscal_month=dt.month,
    )
    db.add(voucher)
    db.flush()
    for e in payload.entries:
        db.add(JournalEntry(
            voucher_id=voucher.id, account_code=e.account_code,
            debit=e.debit, credit=e.credit,
        ))
    db.commit()
    return VoucherCreated(id=voucher.id, voucher_no=v_no)


@app.delete("/api/vouchers/{voucher_id}", status_code=204, tags=["凭证"])
def delete_voucher(
    voucher_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(require_roles("admin", "accountant")),
):
    voucher = db.get(Voucher, voucher_id)
    if voucher is None:
        raise HTTPException(status_code=404, detail="凭证不存在")
    db.delete(voucher)
    db.commit()
    return None


# ── 其他模块（只读列表） ─────────────────────────────
@app.get("/api/inventory/products", tags=["库存"])
def inventory_products(db: Session = Depends(get_db)):
    return repo.list_products(db)


@app.get("/api/inventory/transactions", tags=["库存"])
def inventory_transactions(limit: int = Query(100, ge=1, le=500), db: Session = Depends(get_db)):
    return repo.list_inventory_transactions(db, limit)


@app.get("/api/employees", tags=["薪资"])
def employees(db: Session = Depends(get_db)):
    return repo.list_employees(db)


@app.get("/api/payroll", tags=["薪资"])
def payroll(
    year: int = Query(0, ge=0),
    month: int = Query(0, ge=0, le=12),
    db: Session = Depends(get_db),
):
    return repo.list_payroll_records(db, year or None, month or None)


@app.get("/api/assets", tags=["固定资产"])
def fixed_assets(db: Session = Depends(get_db)):
    return repo.list_fixed_assets(db)


@app.get("/api/projects", tags=["项目"])
def projects(db: Session = Depends(get_db)):
    return repo.list_projects(db)


@app.get("/api/budgets", tags=["预算"])
def budgets(year: int = Query(0, ge=0), db: Session = Depends(get_db)):
    return repo.list_budgets(db, year or None)


@app.get("/api/alerts/rules", tags=["预警"])
def alert_rules(db: Session = Depends(get_db)):
    return repo.list_alert_rules(db)


@app.get("/api/alerts/history", tags=["预警"])
def alert_history(limit: int = Query(100, ge=1, le=500), db: Session = Depends(get_db)):
    return repo.list_alert_history(db, limit)


@app.get("/api/audit", tags=["审计"])
def audit_logs(limit: int = Query(100, ge=1, le=500), db: Session = Depends(get_db)):
    return repo.list_audit_logs(db, limit)


@app.get("/api/esg", tags=["ESG"])
def esg(year: int = Query(0, ge=0), db: Session = Depends(get_db)):
    return repo.list_esg_data(db, year or None)


# ── 写操作（需 accountant / admin 角色） ──────────────
def _commit(db: Session, result):
    db.commit()
    return result


@app.post("/api/inventory/products", status_code=201, tags=["库存"])
def create_product(p: ProductCreate, db: Session = Depends(get_db),
                   user: dict = Depends(require_roles("admin", "accountant"))):
    try:
        r = repo.create_product(db, p.code, p.name, p.category, p.unit,
                                p.unit_price, p.quantity, p.min_stock)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return _commit(db, r)


@app.post("/api/inventory/in", tags=["库存"])
def inventory_in(m: InventoryMove, db: Session = Depends(get_db),
                 user: dict = Depends(require_roles("admin", "accountant"))):
    try:
        r = repo.inventory_in(db, m.product_id, m.quantity, m.unit_price, m.note)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return _commit(db, r)


@app.post("/api/inventory/out", tags=["库存"])
def inventory_out(m: InventoryMove, db: Session = Depends(get_db),
                  user: dict = Depends(require_roles("admin", "accountant"))):
    try:
        r = repo.inventory_out(db, m.product_id, m.quantity, m.note)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return _commit(db, r)


@app.post("/api/employees", status_code=201, tags=["薪资"])
def create_employee(e: EmployeeCreate, db: Session = Depends(get_db),
                    user: dict = Depends(require_roles("admin", "accountant"))):
    try:
        r = repo.add_employee(db, e.code, e.name, e.department, e.position,
                              e.base_salary, e.insurance, e.housing_fund)
    except ValueError as ex:
        raise HTTPException(status_code=400, detail=str(ex))
    return _commit(db, r)


@app.post("/api/payroll/run", tags=["薪资"])
def run_payroll(body: PayrollRun, db: Session = Depends(get_db),
                user: dict = Depends(require_roles("admin", "accountant"))):
    r = repo.calculate_payroll(db, body.year, body.month)
    return _commit(db, {"created": len(r), "records": r})


@app.post("/api/assets", status_code=201, tags=["固定资产"])
def create_asset(a: FixedAssetCreate, db: Session = Depends(get_db),
                 user: dict = Depends(require_roles("admin", "accountant"))):
    try:
        r = repo.add_fixed_asset(db, a.name, a.original_value, a.useful_life_months,
                                 a.purchase_date, a.residual_value, a.depreciation_method)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return _commit(db, r)


@app.post("/api/projects", status_code=201, tags=["项目"])
def create_project(p: ProjectCreate, db: Session = Depends(get_db),
                   user: dict = Depends(require_roles("admin", "accountant"))):
    try:
        r = repo.add_project(db, p.code, p.name, p.budget, p.start_date, p.end_date)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return _commit(db, r)


# ── 凭证明细 / 科目新增 / 试算平衡表 ──────────────────
@app.get("/api/vouchers/{voucher_id}", tags=["凭证"])
def voucher_detail(voucher_id: int, db: Session = Depends(get_db)):
    detail = repo.get_voucher_detail(db, voucher_id)
    if detail is None:
        raise HTTPException(status_code=404, detail="凭证不存在")
    return detail


@app.post("/api/accounts", status_code=201, tags=["科目"])
def create_account(a: AccountCreate, db: Session = Depends(get_db),
                   user: dict = Depends(require_roles("admin", "accountant"))):
    try:
        r = repo.add_account(db, a.code, a.name, a.category, a.nature, a.parent)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return _commit(db, r)


@app.get("/api/reports/trial-balance", tags=["报表"])
def report_trial_balance(db: Session = Depends(get_db)):
    return repo.trial_balance_data(db)


@app.get("/api/trend", tags=["报表"])
def report_trend(year: int = Query(2026), db: Session = Depends(get_db)):
    return repo.trend_data(db, year)


# ── 账龄 / 预算执行 / 项目损益 ────────────────────────
@app.get("/api/aging", tags=["应收应付"])
def aging(db: Session = Depends(get_db)):
    return repo.aging_data(db)


@app.get("/api/budgets/execution", tags=["预算"])
def budget_execution(year: int = Query(2026), db: Session = Depends(get_db)):
    return repo.budget_execution(db, year)


@app.get("/api/projects/pnl", tags=["项目"])
def projects_pnl(db: Session = Depends(get_db)):
    return repo.all_projects_pnl(db)


# ── 编辑操作（需 accountant / admin） ─────────────────
@app.post("/api/accounts/{code}/deactivate", tags=["科目"])
def deactivate_account(code: str, db: Session = Depends(get_db),
                       user: dict = Depends(require_roles("admin", "accountant"))):
    try:
        r = repo.deactivate_account(db, code)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return _commit(db, r)


@app.put("/api/employees/{emp_id}", tags=["薪资"])
def update_employee(emp_id: int, body: EmployeeUpdate, db: Session = Depends(get_db),
                    user: dict = Depends(require_roles("admin", "accountant"))):
    try:
        r = repo.update_employee(db, emp_id, **body.model_dump(exclude_none=True))
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return _commit(db, r)


# ── 报表导出 Excel ───────────────────────────────────
@app.get("/api/reports/export.xlsx", tags=["报表"])
def export_excel(db: Session = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()

    def fill(ws, headers, rows):
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in rows:
            ws.append(r)

    ws1 = wb.active
    ws1.title = "试算平衡表"
    tb = repo.trial_balance_data(db)
    fill(ws1, ["编码", "科目", "借方", "贷方"],
         [[r["code"], r["name"], r["debit"], r["credit"]] for r in tb["rows"]])
    ws1.append(["合计", "", tb["total_debit"], tb["total_credit"]])

    ws2 = wb.create_sheet("资产负债表")
    fill(ws2, ["项目", "金额"], [[r["label"], r["amount"]] for r in repo.balance_sheet_data(db)])

    ws3 = wb.create_sheet("利润表")
    fill(ws3, ["项目", "金额"], [[r["label"], r["amount"]] for r in repo.income_statement_data(db)])

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=report.xlsx"},
    )


# ── AI 自然语言问数 ────────────────────────────────
@app.post("/api/ai/query", response_model=AiQueryResponse, tags=["AI"])
def ai_query(body: AiQueryRequest):
    question = body.question.strip()
    if not question:
        raise HTTPException(status_code=400, detail="问题不能为空")
    return AiQueryResponse(result=acc.ai_query_database(question))
