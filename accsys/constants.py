"""Constants for the accounting system."""

from __future__ import annotations

from typing import Dict, List, Tuple, Any, Optional
from openpyxl.styles import Font, PatternFill, Border, Side

APP_NAME: str = "会计系统专业版"
DB_FILE: str = "accounting.db"
EXCEL_FILE: str = "会计系统报表.xlsx"
ACCOUNTS_FILE: str = "科目表.json"
AI_CONFIG_FILE: str = "ai_config.json"

DEFAULT_ACCOUNTS: Dict[str, Dict[str, Any]] = {
    "1001": {"name": "库存现金", "category": "asset", "nature": "debit", "level": 1, "parent": ""},
    "1002": {"name": "银行存款", "category": "asset", "nature": "debit", "level": 1, "parent": ""},
    "1012": {"name": "其他货币资金", "category": "asset", "nature": "debit", "level": 1, "parent": ""},
    "1101": {"name": "交易性金融资产", "category": "asset", "nature": "debit", "level": 1, "parent": ""},
    "1122": {"name": "应收账款", "category": "asset", "nature": "debit", "level": 1, "parent": ""},
    "1123": {"name": "预付账款", "category": "asset", "nature": "debit", "level": 1, "parent": ""},
    "1221": {"name": "其他应收款", "category": "asset", "nature": "debit", "level": 1, "parent": ""},
    "1403": {"name": "原材料", "category": "asset", "nature": "debit", "level": 1, "parent": ""},
    "1405": {"name": "库存商品", "category": "asset", "nature": "debit", "level": 1, "parent": ""},
    "1601": {"name": "固定资产", "category": "asset", "nature": "debit", "level": 1, "parent": ""},
    "1602": {"name": "累计折旧", "category": "asset", "nature": "credit", "level": 1, "parent": "", "contra": True},
    "1701": {"name": "无形资产", "category": "asset", "nature": "debit", "level": 1, "parent": ""},
    "2001": {"name": "短期借款", "category": "liability", "nature": "credit", "level": 1, "parent": ""},
    "2201": {"name": "应付票据", "category": "liability", "nature": "credit", "level": 1, "parent": ""},
    "2202": {"name": "应付账款", "category": "liability", "nature": "credit", "level": 1, "parent": ""},
    "2203": {"name": "预收账款", "category": "liability", "nature": "credit", "level": 1, "parent": ""},
    "2211": {"name": "应付职工薪酬", "category": "liability", "nature": "credit", "level": 1, "parent": ""},
    "2221": {"name": "应交税费", "category": "liability", "nature": "credit", "level": 1, "parent": ""},
    "2241": {"name": "其他应付款", "category": "liability", "nature": "credit", "level": 1, "parent": ""},
    "2501": {"name": "长期借款", "category": "liability", "nature": "credit", "level": 1, "parent": ""},
    "4001": {"name": "实收资本", "category": "equity", "nature": "credit", "level": 1, "parent": ""},
    "4002": {"name": "资本公积", "category": "equity", "nature": "credit", "level": 1, "parent": ""},
    "4101": {"name": "盈余公积", "category": "equity", "nature": "credit", "level": 1, "parent": ""},
    "4103": {"name": "本年利润", "category": "equity", "nature": "credit", "level": 1, "parent": ""},
    "4104": {"name": "利润分配", "category": "equity", "nature": "credit", "level": 1, "parent": ""},
    "6001": {"name": "主营业务收入", "category": "income", "nature": "credit", "level": 1, "parent": ""},
    "6051": {"name": "其他业务收入", "category": "income", "nature": "credit", "level": 1, "parent": ""},
    "6111": {"name": "投资收益", "category": "income", "nature": "credit", "level": 1, "parent": ""},
    "6301": {"name": "营业外收入", "category": "income", "nature": "credit", "level": 1, "parent": ""},
    "6401": {"name": "主营业务成本", "category": "expense", "nature": "debit", "level": 1, "parent": ""},
    "6402": {"name": "其他业务成本", "category": "expense", "nature": "debit", "level": 1, "parent": ""},
    "6405": {"name": "税金及附加", "category": "expense", "nature": "debit", "level": 1, "parent": ""},
    "6601": {"name": "销售费用", "category": "expense", "nature": "debit", "level": 1, "parent": ""},
    "6602": {"name": "管理费用", "category": "expense", "nature": "debit", "level": 1, "parent": ""},
    "6603": {"name": "财务费用", "category": "expense", "nature": "debit", "level": 1, "parent": ""},
    "6701": {"name": "资产减值损失", "category": "expense", "nature": "debit", "level": 1, "parent": ""},
    "6711": {"name": "营业外支出", "category": "expense", "nature": "debit", "level": 1, "parent": ""},
    "6801": {"name": "所得税费用", "category": "expense", "nature": "debit", "level": 1, "parent": ""},
}

CATEGORY_LABELS: Dict[str, str] = {"asset": "资产类", "liability": "负债类", "equity": "权益类", "income": "收入类", "expense": "费用类"}
NATURE_LABELS: Dict[str, str] = {"debit": "借方余额", "credit": "贷方余额"}
ACCOUNT_CATEGORY_RANGES: Dict[str, str] = {"1": "asset", "2": "liability", "3": "mixed", "4": "equity", "5": "cost", "6": "income/expense"}

EXCHANGE_RATE_API: str = "https://api.exchangerate-api.com/v4/latest/CNY"

PIT_RATES: List[Tuple[int, int, float, int]] = [
    (0, 36000, 0.03, 0),
    (36000, 144000, 0.10, 2520),
    (144000, 300000, 0.20, 16920),
    (300000, 420000, 0.25, 31920),
    (420000, 660000, 0.30, 52920),
    (660000, 960000, 0.35, 85920),
    (960000, float('inf'), 0.45, 181920),
]

FONT_HEADER = Font(bold=True, color="FFFFFF", size=12)
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FILL_ALT = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

VOUCHER_TEMPLATES: Dict[str, Dict[str, Any]] = {
    "提取现金": {
        "summary": "提取备用金",
        "entries": [("1001", "d"), ("1002", "c")],
        "params": [{"name": "提取金额", "key": "amount"}],
        "desc": "从银行提取备用金"
    },
    "支付货款": {
        "summary": "支付供应商货款",
        "entries": [("2202", "d"), ("1002", "c")],
        "params": [{"name": "付款金额", "key": "amount"}],
        "desc": "支付应付账款"
    },
    "收到货款": {
        "summary": "收到客户货款",
        "entries": [("1002", "d"), ("1122", "c")],
        "params": [{"name": "收款金额", "key": "amount"}],
        "desc": "收回应收账款"
    },
    "计提工资": {
        "summary": "计提本月工资",
        "entries": [("6602", "d"), ("2211", "c")],
        "params": [{"name": "工资金额", "key": "amount"}],
        "desc": "计提应付职工薪酬"
    },
    "缴纳税费": {
        "summary": "缴纳增值税及附加",
        "entries": [("2221", "d"), ("1002", "c")],
        "params": [{"name": "纳税金额", "key": "amount"}],
        "desc": "缴纳税费"
    },
    "购买固定资产": {
        "summary": "购入固定资产",
        "entries": [("1601", "d"), ("1002", "c")],
        "params": [{"name": "购买金额", "key": "amount"}],
        "desc": "购置固定资产"
    },
    "收到投资": {
        "summary": "收到股东投资款",
        "entries": [("1002", "d"), ("4001", "c")],
        "params": [{"name": "投资金额", "key": "amount"}],
        "desc": "实收资本增加"
    },
    "偿还借款": {
        "summary": "偿还银行借款",
        "entries": [("2001", "d"), ("1002", "c")],
        "params": [{"name": "还款金额", "key": "amount"}],
        "desc": "偿还短期借款"
    },
}

ESG_INDICATORS: Dict[str, Dict[str, Any]] = {
    "environment": {
        "label": "环境(E)",
        "metrics": [
            ("carbon_emissions", "碳排放量", "tCO2e", "越低越好"),
            ("energy_consumption", "能源消耗量", "MWh", "越低越好"),
            ("water_usage", "用水量", "m³", "越低越好"),
            ("waste_total", "废弃物总量", "吨", "越低越好"),
            ("renewable_energy_pct", "可再生能源占比", "%", "越高越好"),
            ("waste_recycling_rate", "废弃物回收率", "%", "越高越好"),
        ]
    },
    "social": {
        "label": "社会(S)",
        "metrics": [
            ("employee_total", "员工总数", "人", "适中为好"),
            ("employee_female_pct", "女性员工比例", "%", "越高越好"),
            ("training_hours_per_employee", "人均培训时长", "小时/年", "越高越好"),
            ("safety_incidents", "安全事故数", "次", "越低越好"),
            ("community_investment", "社区投资额", "元", "越高越好"),
            ("employee_turnover_pct", "员工流失率", "%", "越低越好"),
            ("customer_satisfaction", "客户满意度", "分(0-100)", "越高越好"),
        ]
    },
    "governance": {
        "label": "治理(G)",
        "metrics": [
            ("board_meetings", "董事会会议次数", "次/年", "适中为好"),
            ("board_independent_pct", "独立董事比例", "%", "越高越好"),
            ("compliance_violations", "合规违规事件数", "次", "越低越好"),
            ("audit_findings", "审计发现数", "项", "越低越好"),
            ("ethics_training_pct", "道德培训覆盖率", "%", "越高越好"),
            ("data_breaches", "数据泄露事件数", "次", "越低越好"),
        ]
    }
}

ESG_SCORE_DIRECTION: Dict[str, Optional[bool]] = {
    "carbon_emissions": False, "energy_consumption": False, "water_usage": False,
    "waste_total": False, "renewable_energy_pct": True, "waste_recycling_rate": True,
    "employee_total": None, "employee_female_pct": True,
    "training_hours_per_employee": True, "safety_incidents": False,
    "community_investment": True, "employee_turnover_pct": False,
    "customer_satisfaction": True,
    "board_meetings": None, "board_independent_pct": True,
    "compliance_violations": False, "audit_findings": False,
    "ethics_training_pct": True, "data_breaches": False,
}

STUDENT_TAX_DEDUCTIONS: List[Tuple[str, str, int]] = [
    ("基本减除费用", "每月5000元，全年60000元", 60000),
    ("专项附加扣除-教育", "继续教育每月400元", 4800),
    ("专项附加扣除-租房", "根据城市每月800-1500元", 12000),
    ("专项附加扣除-大病医疗", "年度超过15000元部分据实扣除", 0),
]

STUDENT_TAX_TIPS: List[str] = [
    '* 大学生兼职收入属于劳务报酬所得，按次或按月预扣预缴',
    '* 年度汇算清缴时，劳务报酬与工资薪金合并计算综合所得',
    '* 年收入不超过6万元可申请退税（已预缴的税款）',
    '* 通过个人所得税APP在每年3月-6月办理汇算清缴',
    '* 保留兼职合同、银行流水等凭证备查',
    '* 多个平台兼职的收入需要合并申报',
]

MICRO_CATEGORIES: Dict[str, List[str]] = {
    "income": ["销售收入", "服务收入", "投资收益", "其他收入"],
    "expense": ["采购成本", "房租水电", "员工工资", "交通差旅", "餐饮招待",
                 "办公用品", "市场推广", "设备维护", "税费", "其他支出"],
}

CERT_EXAMS: Dict[str, Dict[str, Any]] = {
    "primary_accounting": {
        "name": "初级会计职称",
        "subjects": ["初级会计实务", "经济法基础"],
        "typical_date": "5月",
        "registration": "前一年11月-当年1月",
        "study_hours": {"初级会计实务": 120, "经济法基础": 100},
        "total_hours": 220,
        "difficulty": "★★☆☆☆",
        "description": "会计入门证书，适合在校生和转行者",
    },
    "cpa": {
        "name": "CPA注册会计师",
        "subjects": ["会计", "审计", "财务成本管理", "税法", "经济法", "公司战略与风险管理"],
        "typical_date": "8月",
        "registration": "4月",
        "study_hours": {"会计": 400, "审计": 350, "财务成本管理": 300, "税法": 250, "经济法": 200, "公司战略与风险管理": 200},
        "total_hours": 1700,
        "difficulty": "★★★★★",
        "description": "会计顶级证书，6门专业课+综合阶段",
    },
    "intermediate_accounting": {
        "name": "中级会计职称",
        "subjects": ["中级会计实务", "财务管理", "经济法"],
        "typical_date": "9月",
        "registration": "3月",
        "study_hours": {"中级会计实务": 200, "财务管理": 150, "经济法": 120},
        "total_hours": 470,
        "difficulty": "★★★☆☆",
        "description": "会计进阶证书，需满足工作年限要求",
    },
    "tax_agent": {
        "name": "税务师",
        "subjects": ["税法一", "税法二", "涉税服务实务", "涉税服务相关法律", "财务与会计"],
        "typical_date": "11月",
        "registration": "5月-7月",
        "study_hours": {"税法一": 150, "税法二": 150, "涉税服务实务": 200, "涉税服务相关法律": 180, "财务与会计": 180},
        "total_hours": 860,
        "difficulty": "★★★★☆",
        "description": "税务领域专业证书，与CPA税法互补",
    },
}

DEFAULT_AI_CONFIG: Dict[str, Any] = {
    "enabled": False,
    "provider": "openai",
    "endpoint": "https://api.openai.com/v1/chat/completions",
    "api_key": "",
    "model": "gpt-3.5-turbo",
    "max_tokens": 2048,
    "temperature": 0.7,
}

THEME_FILE: str = "theme.json"
THEMES: Dict[str, Dict[str, str]] = {
    "light": {
        "bg": "#FFF0F5", "primary": "#FF6B9D", "success": "#7BC67E",
        "warn": "#FFB347", "alt": "#F5E6FF", "accent": "#8B5CF6",
        "violet": "#7C3AED", "purple": "#9C27B0", "fg": "#333333",
        "text_bg": "#FFF0F5", "header_fg": "white", "tab_bg": "#FFE4EC",
        "tab_fg": "#6B3FA0", "tab_sel": "#FFB3C6", "tab_act": "#FFCDD9",
        "tab_sel_fg": "#4A1A6B",
    },
    "dark": {
        "bg": "#1A1A2E", "primary": "#E040A0", "success": "#66BB6A",
        "warn": "#FFA726", "alt": "#2D2D44", "accent": "#B388FF",
        "violet": "#9C7CF4", "purple": "#CE93D8", "fg": "#E0E0E0",
        "text_bg": "#252540", "header_fg": "white", "tab_bg": "#2D2D44",
        "tab_fg": "#B0B0FF", "tab_sel": "#4A1A6B", "tab_act": "#3D2D5C",
        "tab_sel_fg": "#E0E0FF",
    },
}

AR_ACCOUNTS: Dict[str, str] = {"1122": "应收账款", "1123": "预付账款", "1221": "其他应收款"}
AP_ACCOUNTS: Dict[str, str] = {"2202": "应付账款", "2201": "应付票据", "2241": "其他应付款"}

BAD_DEBT_RATES: List[Tuple[int, float]] = [(30, 0.005), (60, 0.01), (90, 0.03), (180, 0.10), (365, 0.30), (float('inf'), 0.50)]

RAG_SYSTEM_PROMPT: str = """你是会计系统的AI查询助手。数据库包含：
- accounts: 科目表(code,name,category,balance)
- vouchers: 凭证(id,date,summary,fiscal_year,fiscal_month)
- journal_entries: 分录(voucher_id,account_code,debit,credit)
- fixed_assets: 固定资产
- budgets: 预算
- products: 库存商品
- inventory_transactions: 库存流水

用户会问中文财务问题。请生成 SQLite SQL 查询语句回答。
只返回 SQL，不要额外解释。SQL 要用聚合函数和过滤条件。
如果问题不明确，返回: --NEED_CLARIFY--"""

DEFAULT_ALERT_RULES: List[Tuple[str, str, str, float, str]] = [
    ("流动比率过低", "current_ratio", "lt", 1.5, "warning"),
    ("资产负债率过高", "debt_ratio", "gt", 0.7, "critical"),
    ("净利润率为负", "net_profit_margin", "lt", 0, "critical"),
    ("速动比率过低", "quick_ratio", "lt", 1.0, "warning"),
    ("毛利率过低", "gross_margin", "lt", 0.15, "warning"),
]

ATTACHMENT_DIR: str = "attachments"
