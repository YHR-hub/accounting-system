"""
报表与分析模块 (Reports)

提供三张报表（资产负债表、利润表、现金流量表）、Excel 导出、
财务比率计算、异常检测、智能建议及趋势预测等功能。
"""

from __future__ import annotations

import sqlite3
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .database import get_conn
from .accounts import load_accounts_from_db, get_account_dict, calc_balances
from .vouchers import next_voucher_no


FONT_HEADER = Font(bold=True, color="FFFFFF", size=12)
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FILL_ALT = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))


def balance_sheet() -> None:
    accounts = load_accounts_from_db()
    balances = calc_balances(accounts)

    asset_codes = [a['code'] for a in accounts if a['category'] == 'asset']
    liability_codes = [a['code'] for a in accounts if a['category'] == 'liability']
    equity_codes = [a['code'] for a in accounts if a['category'] == 'equity']

    assets: List[Tuple[str, str, Decimal]] = []
    liabilities: List[Tuple[str, str, Decimal]] = []
    equities: List[Tuple[str, str, Decimal]] = []

    for a in accounts:
        code = a['code']
        bal = balances.get(code, Decimal('0'))
        if a['category'] == 'asset':
            if a['is_contra']:
                assets.append((code, a['name'], -bal))
            else:
                assets.append((code, a['name'], bal))
        elif a['category'] == 'liability':
            liabilities.append((code, a['name'], bal))
        elif a['category'] == 'equity':
            equities.append((code, a['name'], bal))

    income_acc = [a for a in accounts if a['category'] == 'income']
    expense_acc = [a for a in accounts if a['category'] == 'expense']
    total_income = sum(balances.get(a['code'], Decimal('0')) for a in income_acc)
    total_expense = sum(balances.get(a['code'], Decimal('0')) for a in expense_acc)
    net_profit = total_income - total_expense

    print(f"\n{'=' * 64}")
    print(f"  资产负债表 (Balance Sheet)")
    print(f"  日期: {datetime.now().strftime('%Y-%m-%d')}")
    print(f"{'=' * 64}")
    print(f"{'资产':<32} {'负债及所有者权益':<32}")
    print("-" * 64)

    total_asset = sum(b for _, _, b in assets) + sum(-b for _, _, b in assets if b < 0)
    total_asset = sum(abs(b) for _, _, b in assets)
    total_liability = sum(b for _, _, b in liabilities)
    total_equity = sum(b for _, _, b in equities) + max(net_profit, Decimal('0'))

    max_rows = max(len(assets), len(liabilities) + len(equities) + 1)
    for i in range(max_rows):
        left = ""
        right = ""
        if i < len(assets):
            c, n, b = assets[i]
            sign = "" if not any(x['is_contra'] for x in accounts if x['code'] == c) else "(-)"
            left = f"{n:<22} {float(b):>10.2f}"
        if i < len(liabilities):
            c, n, b = liabilities[i]
            right = f"{n:<22} {float(b):>10.2f}"
        elif i == len(liabilities):
            right = f"{'净利润(本年)':<22} {float(net_profit):>10.2f}"
        elif i < len(liabilities) + len(equities) + 1:
            idx = i - len(liabilities) - 1
            if idx < len(equities):
                c, n, b = equities[idx]
                right = f"{n:<22} {float(b):>10.2f}"

        print(f"{left:<32} {right:<32}")

    print("-" * 64)
    net_equity = total_equity
    print(f"{'资产总计':<22} {float(total_asset):>10.2f} {'负债及权益总计':<22} {float(total_liability + net_equity):>10.2f}")
    if abs(total_asset - total_liability - net_equity) < Decimal('0.01'):
        print(f"  [OK] 会计恒等式成立: 资产 = 负债 + 所有者权益")
    else:
        diff = total_asset - total_liability - net_equity
        print(f"  [ERR] 不平! 差额: {float(diff):.2f}")
    print(f"{'=' * 64}")


def income_statement() -> None:
    accounts = load_accounts_from_db()
    balances = calc_balances(accounts)

    income_acc = [a for a in accounts if a['category'] == 'income']
    expense_acc = [a for a in accounts if a['category'] == 'expense']

    print(f"\n{'=' * 56}")
    print(f"  利润表 (Income Statement)")
    print(f"  期间: {datetime.now().strftime('%Y-%m')}")
    print(f"{'=' * 56}")
    print(f"{'项目':<30} {'金额':>12}")
    print("-" * 56)

    print(f"{'一、营业收入':<30}")
    operating_revenue = Decimal('0')
    for a in income_acc:
        if '收入' in a['name']:
            bal = balances.get(a['code'], Decimal('0'))
            operating_revenue += bal
            print(f"  {a['name']:<26} {float(bal):>12.2f}")
    print(f"{'  营业收入合计':<28} {float(operating_revenue):>12.2f}")

    print()
    print(f"{'二、营业成本及费用':<30}")
    operating_expense = Decimal('0')
    for a in expense_acc:
        if a['code'] not in ('6801',):
            bal = balances.get(a['code'], Decimal('0'))
            operating_expense += bal
            print(f"  {a['name']:<26} {float(bal):>12.2f}")
    print(f"{'  营业成本合计':<28} {float(operating_expense):>12.2f}")

    gross_profit = operating_revenue - operating_expense
    print(f"\n{'三、营业利润':<28} {float(gross_profit):>12.2f}")

    other_income = Decimal('0')
    for a in income_acc:
        if '收入' not in a['name']:
            bal = balances.get(a['code'], Decimal('0'))
            other_income += bal
            print(f"  {a['name']:<26} {float(bal):>12.2f}")
    if other_income != 0:
        print(f"{'  其他收益合计':<28} {float(other_income):>12.2f}")

    tax_exp = balances.get('6801', Decimal('0')) if any(a['code'] == '6801' for a in expense_acc) else Decimal('0')
    if tax_exp > 0:
        print(f"\n  减: {next(a['name'] for a in expense_acc if a['code'] == '6801'):<20} {float(tax_exp):>12.2f}")

    net_profit = gross_profit + other_income - tax_exp
    print(f"\n{'四、净利润':<28} {float(net_profit):>12.2f}")
    print(f"{'=' * 56}")


def cash_flow_statement() -> None:
    accounts = load_accounts_from_db()
    acc_dict = get_account_dict(accounts)

    conn = get_conn()
    rows = conn.execute("""
        SELECT v.date, v.summary, je.account_code, je.debit, je.credit
        FROM journal_entries je
        JOIN vouchers v ON je.voucher_id = v.id
        ORDER BY v.date
    """).fetchall()
    conn.close()

    operating_inflow = Decimal('0')
    operating_outflow = Decimal('0')
    investing_inflow = Decimal('0')
    investing_outflow = Decimal('0')
    financing_inflow = Decimal('0')
    financing_outflow = Decimal('0')

    for r in rows:
        code = r['account_code']
        debit = Decimal(str(r['debit']))
        credit = Decimal(str(r['credit']))
        a = acc_dict.get(code)
        if not a:
            continue
        cat = a['category']

        if cat == 'income':
            operating_inflow += credit
        elif cat == 'expense':
            if code in ('6602', '6601', '6401', '6402', '6405'):
                operating_outflow += debit
            elif code == '6603':
                operating_outflow += debit
        elif code == '1002':
            operating_inflow += credit
            operating_outflow += debit
        elif code == '1601':
            investing_outflow += debit
            investing_inflow += credit
        elif code in ('2001', '2501'):
            financing_inflow += credit
            financing_outflow += debit
        elif code == '4001':
            financing_inflow += credit

    operating_net = operating_inflow - operating_outflow
    investing_net = investing_inflow - investing_outflow
    financing_net = financing_inflow - financing_outflow
    net_change = operating_net + investing_net + financing_net

    print(f"\n{'=' * 56}")
    print(f"  现金流量表 (Cash Flow Statement)")
    print(f"  期间: {datetime.now().strftime('%Y-%m')}")
    print(f"{'=' * 56}")
    print(f"{'项目':<32} {'金额':>12}")
    print("-" * 56)
    print(f"{'一、经营活动现金流':<32}")
    print(f"  流入:{'':<24} {float(operating_inflow):>12.2f}")
    print(f"  流出:{'':<24} {float(operating_outflow):>12.2f}")
    print(f"  净额:{'':<24} {float(operating_net):>12.2f}")
    print()
    print(f"{'二、投资活动现金流':<32}")
    print(f"  流入:{'':<24} {float(investing_inflow):>12.2f}")
    print(f"  流出:{'':<24} {float(investing_outflow):>12.2f}")
    print(f"  净额:{'':<24} {float(investing_net):>12.2f}")
    print()
    print(f"{'三、筹资活动现金流':<32}")
    print(f"  流入:{'':<24} {float(financing_inflow):>12.2f}")
    print(f"  流出:{'':<24} {float(financing_outflow):>12.2f}")
    print(f"  净额:{'':<24} {float(financing_net):>12.2f}")
    print("-" * 56)
    print(f"{'四、现金净增加额':<30} {float(net_change):>12.2f}")
    print(f"{'=' * 56}")


def balance_sheet_data() -> List[Dict[str, Any]]:
    """资产负债表结构化数据（供 Web/API 使用），返回 [{label, amount}] 行列表。"""
    accounts = load_accounts_from_db()
    balances = calc_balances(accounts)

    assets: List[Tuple[str, Decimal]] = []
    liabilities: List[Tuple[str, Decimal]] = []
    equities: List[Tuple[str, Decimal]] = []
    for a in accounts:
        bal = balances.get(a['code'], Decimal('0'))
        if a['category'] == 'asset':
            assets.append((a['name'], -bal if a['is_contra'] else bal))
        elif a['category'] == 'liability':
            liabilities.append((a['name'], bal))
        elif a['category'] == 'equity':
            equities.append((a['name'], bal))

    income_acc = [a for a in accounts if a['category'] == 'income']
    expense_acc = [a for a in accounts if a['category'] == 'expense']
    total_income = sum(balances.get(a['code'], Decimal('0')) for a in income_acc)
    total_expense = sum(balances.get(a['code'], Decimal('0')) for a in expense_acc)
    net_profit = total_income - total_expense

    total_asset = sum((abs(b) for _, b in assets), Decimal('0'))
    total_liability = sum((b for _, b in liabilities), Decimal('0'))
    total_equity = sum((b for _, b in equities), Decimal('0')) + max(net_profit, Decimal('0'))

    rows: List[Dict[str, Any]] = []
    for name, b in assets:
        rows.append({'label': name, 'amount': float(b), 'section': 'asset'})
    rows.append({'label': '资产总计', 'amount': float(total_asset), 'section': 'total'})
    for name, b in liabilities:
        rows.append({'label': name, 'amount': float(b), 'section': 'liability'})
    rows.append({'label': '净利润(本年)', 'amount': float(net_profit), 'section': 'equity'})
    for name, b in equities:
        rows.append({'label': name, 'amount': float(b), 'section': 'equity'})
    rows.append({'label': '负债及所有者权益总计',
                 'amount': float(total_liability + total_equity), 'section': 'total'})
    return rows


def income_statement_data() -> List[Dict[str, Any]]:
    """利润表结构化数据，返回 [{label, amount}] 行列表。"""
    accounts = load_accounts_from_db()
    balances = calc_balances(accounts)
    income_acc = [a for a in accounts if a['category'] == 'income']
    expense_acc = [a for a in accounts if a['category'] == 'expense']

    operating_revenue = Decimal('0')
    other_income = Decimal('0')
    for a in income_acc:
        bal = balances.get(a['code'], Decimal('0'))
        if '收入' in a['name']:
            operating_revenue += bal
        else:
            other_income += bal

    operating_expense = sum(
        (balances.get(a['code'], Decimal('0')) for a in expense_acc if a['code'] not in ('6801',)),
        Decimal('0'),
    )
    tax_exp = balances.get('6801', Decimal('0')) if any(a['code'] == '6801' for a in expense_acc) else Decimal('0')

    gross_profit = operating_revenue - operating_expense
    net_profit = gross_profit + other_income - tax_exp

    rows: List[Dict[str, Any]] = [
        {'label': '一、营业收入', 'amount': float(operating_revenue)},
        {'label': '二、营业成本及费用', 'amount': float(operating_expense)},
        {'label': '三、营业利润', 'amount': float(gross_profit)},
    ]
    if other_income != 0:
        rows.append({'label': '加：其他收益', 'amount': float(other_income)})
    if tax_exp > 0:
        rows.append({'label': '减：所得税费用', 'amount': float(tax_exp)})
    rows.append({'label': '四、净利润', 'amount': float(net_profit)})
    return rows


def cash_flow_statement_data() -> List[Dict[str, Any]]:
    """现金流量表（间接法）结构化数据，返回 [{label, amount}] 行列表。"""
    accounts = load_accounts_from_db()
    acc_dict = get_account_dict(accounts)

    conn = get_conn()
    rows_db = conn.execute("""
        SELECT v.date, v.summary, je.account_code, je.debit, je.credit
        FROM journal_entries je
        JOIN vouchers v ON je.voucher_id = v.id
        ORDER BY v.date
    """).fetchall()
    conn.close()

    operating_inflow = operating_outflow = Decimal('0')
    investing_inflow = investing_outflow = Decimal('0')
    financing_inflow = financing_outflow = Decimal('0')

    for r in rows_db:
        code = r['account_code']
        debit = Decimal(str(r['debit']))
        credit = Decimal(str(r['credit']))
        a = acc_dict.get(code)
        if not a:
            continue
        cat = a['category']
        if cat == 'income':
            operating_inflow += credit
        elif cat == 'expense':
            if code in ('6602', '6601', '6401', '6402', '6405', '6603'):
                operating_outflow += debit
        elif code == '1002':
            operating_inflow += credit
            operating_outflow += debit
        elif code == '1601':
            investing_outflow += debit
            investing_inflow += credit
        elif code in ('2001', '2501'):
            financing_inflow += credit
            financing_outflow += debit
        elif code == '4001':
            financing_inflow += credit

    operating_net = operating_inflow - operating_outflow
    investing_net = investing_inflow - investing_outflow
    financing_net = financing_inflow - financing_outflow
    net_change = operating_net + investing_net + financing_net

    return [
        {'label': '一、经营活动现金流量净额', 'amount': float(operating_net)},
        {'label': '二、投资活动现金流量净额', 'amount': float(investing_net)},
        {'label': '三、筹资活动现金流量净额', 'amount': float(financing_net)},
        {'label': '四、现金及现金等价物净增加额', 'amount': float(net_change)},
    ]


def export_excel() -> None:
    accounts = load_accounts_from_db()
    balances = calc_balances(accounts)

    wb = openpyxl.Workbook()

    _excel_trial_balance(wb, accounts, balances)
    _excel_balance_sheet(wb, accounts, balances)
    _excel_income_statement(wb, accounts, balances)

    from .database import get_excel_path
    path = get_excel_path()
    wb.save(path)
    print(f"\n[OK] 报表已导出: {path}")


def _style_header(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _style_row(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, max_col: int, alt: bool = False) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="right" if col > 1 else "left")
        if alt:
            cell.fill = FILL_ALT


def _excel_trial_balance(wb: openpyxl.Workbook, accounts: List[dict], balances: Dict[str, Decimal]) -> None:
    ws = wb.active
    ws.title = "试算平衡表"
    ws.cell(row=1, column=1, value="试算平衡表").font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    headers = ["科目编码", "科目名称", "借方余额", "贷方余额"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    _style_header(ws, 2, 4)

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16

    row = 3
    total_dr = Decimal('0')
    total_cr = Decimal('0')
    for a in accounts:
        code = a['code']
        bal = balances.get(code, Decimal('0'))
        dr = Decimal('0')
        cr = Decimal('0')
        if bal > 0:
            if a['nature'] == 'debit':
                dr = bal
            else:
                cr = bal
        elif bal < 0:
            if a['nature'] == 'debit':
                cr = -bal
            else:
                dr = -bal

        ws.cell(row=row, column=1, value=code)
        ws.cell(row=row, column=2, value=a['name'])
        ws.cell(row=row, column=3, value=float(dr))
        ws.cell(row=row, column=4, value=float(cr))
        _style_row(ws, row, 4, alt=(row % 2 == 0))
        total_dr += dr
        total_cr += cr
        row += 1

    ws.cell(row=row, column=1, value="合计").font = Font(bold=True)
    ws.cell(row=row, column=2, value="")
    ws.cell(row=row, column=3, value=float(total_dr)).font = Font(bold=True)
    ws.cell(row=row, column=4, value=float(total_cr)).font = Font(bold=True)
    for c in range(1, 5):
        ws.cell(row=row, column=c).border = THIN_BORDER


def _excel_balance_sheet(wb: openpyxl.Workbook, accounts: List[dict], balances: Dict[str, Decimal]) -> None:
    ws = wb.create_sheet("资产负债表")
    ws.cell(row=1, column=1, value="资产负债表").font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16

    for i, h in enumerate(["资产", "金额", "负债及权益", "金额"], 1):
        ws.cell(row=2, column=i, value=h)
    _style_header(ws, 2, 4)

    assets = [(a['code'], a['name'], balances.get(a['code'], Decimal('0')))
              for a in accounts if a['category'] == 'asset' and not a['is_contra']]
    contra_assets = [(a['code'], a['name'], balances.get(a['code'], Decimal('0')))
                     for a in accounts if a['category'] == 'asset' and a['is_contra']]
    liabilities = [(a['code'], a['name'], balances.get(a['code'], Decimal('0')))
                   for a in accounts if a['category'] == 'liability']
    equities = [(a['code'], a['name'], balances.get(a['code'], Decimal('0')))
                for a in accounts if a['category'] == 'equity']

    income_acc = [a for a in accounts if a['category'] == 'income']
    expense_acc = [a for a in accounts if a['category'] == 'expense']
    total_income = sum(balances.get(a['code'], Decimal('0')) for a in income_acc)
    total_expense = sum(balances.get(a['code'], Decimal('0')) for a in expense_acc)
    net_profit = max(total_income - total_expense, Decimal('0'))

    max_rows = max(len(assets) + len(contra_assets) + 1,
                   len(liabilities) + len(equities) + 2)
    row = 3
    for i in range(max_rows):
        left_code, left_name, left_amt = "", "", Decimal('0')
        right_code, right_name, right_amt = "", "", Decimal('0')

        if i < len(assets):
            left_code, left_name, left_amt = assets[i]
        elif i < len(assets) + len(contra_assets):
            idx = i - len(assets)
            left_code, left_name, left_amt = contra_assets[idx]
            left_amt = -left_amt
            left_name = "减: " + left_name

        if i < len(liabilities):
            right_code, right_name, right_amt = liabilities[i]
        elif i == len(liabilities):
            right_name = "净利润(本年)"
            right_amt = net_profit
        elif i < len(liabilities) + len(equities) + 1:
            idx = i - len(liabilities) - 1
            if idx < len(equities):
                right_code, right_name, right_amt = equities[idx]

        ws.cell(row=row, column=1, value=left_name)
        ws.cell(row=row, column=2, value=float(left_amt) if left_amt != 0 else "")
        ws.cell(row=row, column=3, value=right_name)
        ws.cell(row=row, column=4, value=float(right_amt) if right_amt != 0 else "")
        _style_row(ws, row, 4, alt=(row % 2 == 0))
        row += 1


def _excel_income_statement(wb: openpyxl.Workbook, accounts: List[dict], balances: Dict[str, Decimal]) -> None:
    ws = wb.create_sheet("利润表")
    ws.cell(row=1, column=1, value="利润表").font = Font(bold=True, size=14)
    ws.merge_cells("A1:C1")

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14

    for i, h in enumerate(["项目", "行次", "金额"], 1):
        ws.cell(row=2, column=i, value=h)
    _style_header(ws, 2, 3)

    income_acc = [a for a in accounts if a['category'] == 'income']
    expense_acc = [a for a in accounts if a['category'] == 'expense']

    row = 3
    ws.cell(row=row, column=1, value="一、营业收入").font = Font(bold=True)
    ws.cell(row=row, column=2, value=1)
    _style_row(ws, row, 3)
    row += 1
    for a in income_acc:
        bal = balances.get(a['code'], Decimal('0'))
        if bal > 0:
            ws.cell(row=row, column=1, value=f"  {a['name']}")
            ws.cell(row=row, column=2, value="")
            ws.cell(row=row, column=3, value=float(bal))
            _style_row(ws, row, 3, alt=True)
            row += 1

    operating_revenue = sum(balances.get(a['code'], Decimal('0')) for a in income_acc)
    ws.cell(row=row, column=1, value="  营业收入合计").font = Font(bold=True)
    ws.cell(row=row, column=3, value=float(operating_revenue)).font = Font(bold=True)
    _style_row(ws, row, 3)
    row += 2

    ws.cell(row=row, column=1, value="二、营业成本及费用").font = Font(bold=True)
    ws.cell(row=row, column=2, value=2)
    _style_row(ws, row, 3)
    row += 1
    operating_expense = Decimal('0')
    for a in expense_acc:
        bal = balances.get(a['code'], Decimal('0'))
        if bal > 0:
            ws.cell(row=row, column=1, value=f"  {a['name']}")
            ws.cell(row=row, column=3, value=float(bal))
            _style_row(ws, row, 3, alt=True)
            if a['code'] != '6801':
                operating_expense += bal
            row += 1

    gross_profit = operating_revenue - operating_expense
    tax_exp = balances.get('6801', Decimal('0'))

    ws.cell(row=row, column=1, value="三、营业利润").font = Font(bold=True)
    ws.cell(row=row, column=2, value=3)
    ws.cell(row=row, column=3, value=float(gross_profit)).font = Font(bold=True)
    _style_row(ws, row, 3)
    row += 1
    if tax_exp > 0:
        ws.cell(row=row, column=1, value=f"  减:所得税费用")
        ws.cell(row=row, column=3, value=float(tax_exp))
        _style_row(ws, row, 3, alt=True)
        row += 1

    net_profit = gross_profit - tax_exp
    ws.cell(row=row, column=1, value="四、净利润").font = Font(bold=True, size=12)
    ws.cell(row=row, column=2, value=4)
    ws.cell(row=row, column=3, value=float(net_profit)).font = Font(bold=True, size=12)
    _style_row(ws, row, 3)


def calc_financial_ratios() -> Dict[str, float]:
    accounts = load_accounts_from_db()
    balances = calc_balances(accounts)
    acc_dict = get_account_dict(accounts)

    asset_total = Decimal('0')
    liability_total = Decimal('0')
    equity_total = Decimal('0')
    current_asset = Decimal('0')
    current_liability = Decimal('0')
    inventory = Decimal('0')
    receivable = Decimal('0')
    revenue = Decimal('0')
    cost = Decimal('0')
    expense_total = Decimal('0')
    cash = Decimal('0')

    for a in accounts:
        code = a['code']
        bal = balances.get(code, Decimal('0'))
        cat = a['category']

        if cat == 'asset':
            asset_total += bal
            if code in ('1001', '1002', '1012', '1101'):
                current_asset += bal
            if code in ('1122',):
                receivable += bal
            if code in ('1403', '1405'):
                inventory += bal
        elif cat == 'liability':
            liability_total += bal
            if code in ('2001', '2201', '2202', '2203', '2211', '2221', '2241'):
                current_liability += bal
        elif cat == 'equity':
            equity_total += bal
        elif cat == 'income':
            revenue += bal
        elif cat == 'expense':
            expense_total += bal
            if code == '6401':
                cost += bal

        if code in ('1001', '1002'):
            cash += bal

    ratios: Dict[str, float] = {}

    current_asset_f = float(current_asset)
    current_liability_f = float(current_liability)
    ratios["流动比率"] = round(current_asset_f / current_liability_f, 2) if current_liability_f else 0
    quick_asset = current_asset_f - float(inventory)
    ratios["速动比率"] = round(quick_asset / current_liability_f, 2) if current_liability_f else 0

    asset_total_f = float(asset_total)
    liability_total_f = float(liability_total)
    ratios["资产负债率"] = round(liability_total_f / asset_total_f * 100, 2) if asset_total_f else 0

    revenue_f = float(revenue)
    cost_f = float(cost)
    ratios["毛利率"] = round((revenue_f - cost_f) / revenue_f * 100, 2) if revenue_f else 0

    expense_total_f = float(expense_total)
    net_profit = revenue_f - expense_total_f
    ratios["净利率"] = round(net_profit / revenue_f * 100, 2) if revenue_f else 0

    equity_total_f = float(equity_total)
    ratios["净资产收益率(ROE)"] = round(net_profit / equity_total_f * 100, 2) if equity_total_f else 0
    ratios["总资产报酬率(ROA)"] = round(net_profit / asset_total_f * 100, 2) if asset_total_f else 0

    ratios["应收账款周转率"] = round(revenue_f / float(receivable), 2) if float(receivable) else 0
    ratios["存货周转率"] = round(cost_f / float(inventory), 2) if float(inventory) else 0

    profit_ratio = net_profit / revenue_f * 100 if revenue_f else 0
    expense_ratio = expense_total_f / revenue_f * 100 if revenue_f else 0
    ratios["费用占收入比"] = round(expense_ratio, 2)
    ratios["净利润率"] = round(profit_ratio, 2)

    return ratios


def detect_anomalies() -> List[Dict[str, Any]]:
    conn = get_conn()
    rows = conn.execute("""
        SELECT je.*, a.name as account_name, v.date, v.summary, v.voucher_no
        FROM journal_entries je
        JOIN accounts a ON je.account_code = a.code
        JOIN vouchers v ON je.voucher_id = v.id
        ORDER BY v.date DESC
    """).fetchall()
    conn.close()

    anomalies = []
    amounts = [max(r['debit'], r['credit']) for r in rows if max(r['debit'], r['credit']) > 0]
    if not amounts:
        return anomalies

    avg = sum(amounts) / len(amounts)
    variance = sum((x - avg) ** 2 for x in amounts) / len(amounts)
    std_dev = variance ** 0.5
    threshold = avg + 3 * std_dev

    for r in rows:
        amt = max(r['debit'], r['credit'])
        if amt > threshold:
            anomalies.append({
                "type": "大额异常",
                "voucher_no": r['voucher_no'],
                "date": r['date'],
                "summary": r['summary'],
                "account": r['account_name'],
                "amount": amt,
                "threshold": round(threshold, 2),
                "detail": f"金额 {amt:.2f} 超过阈值 {threshold:.2f}"
            })

    seen_codes = set()
    for r in rows:
        code = r['account_code']
        if code not in seen_codes:
            seen_codes.add(code)
            total_debit = 0
            total_credit = 0
            conn2 = get_conn()
            stats = conn2.execute(
                "SELECT SUM(debit) as d, SUM(credit) as c, COUNT(*) as cnt FROM journal_entries WHERE account_code=?",
                (code,)).fetchone()
            conn2.close()
            if stats and stats['cnt'] > 5:
                avg_amt = (stats['d'] + stats['c']) / stats['cnt']
                for r2 in rows:
                    if r2['account_code'] == code:
                        amt2 = max(r2['debit'], r2['credit'])
                        if amt2 > 0 and amt2 > avg_amt * 3 and amt2 > 10000:
                            anomalies.append({
                                "type": "科目异常波动",
                                "voucher_no": r2['voucher_no'],
                                "date": r2['date'],
                                "summary": r2['summary'],
                                "account": r2['account_name'],
                                "amount": amt2,
                                "threshold": round(avg_amt * 3, 2),
                                "detail": f"科目 {r2['account_name']} 该笔 {amt2:.2f} 远超平均 {avg_amt:.2f}"
                            })
                            break

    return anomalies


def get_smart_suggestions(ratios: Optional[Dict[str, float]] = None) -> List[str]:
    if ratios is None:
        ratios = calc_financial_ratios()

    suggestions = []

    cr = ratios.get("流动比率", 0)
    if cr == 0:
        suggestions.append("💡 暂无足够数据生成财务建议，请先录入凭证。")
        return suggestions
    elif cr < 1:
        suggestions.append("⚠️ 流动比率偏低 ({:.2f})，短期偿债压力较大，建议增加流动资产或减少短期负债。".format(cr))
    elif cr < 2:
        suggestions.append("📊 流动比率 ({:.2f}) 处于合理偏低范围，建议适当补充流动资金。".format(cr))
    else:
        suggestions.append("✅ 流动比率 ({:.2f}) 良好，短期偿债能力强。".format(cr))

    qr = ratios.get("速动比率", 0)
    if 0 < qr < 0.5:
        suggestions.append("⚠️ 速动比率偏低 ({:.2f})，扣除存货后偿债能力不足，需关注存货积压问题。".format(qr))
    elif qr >= 1:
        suggestions.append("✅ 速动比率 ({:.2f}) 良好，即时偿债能力强。".format(qr))

    dar = ratios.get("资产负债率", 0)
    if dar > 80:
        suggestions.append("🔴 资产负债率偏高 ({:.2f}%)，财务杠杆过大，建议控制负债规模。".format(dar))
    elif dar > 60:
        suggestions.append("🟡 资产负债率 ({:.2f}%) 处于警戒区间，关注债务结构。".format(dar))
    elif dar > 0:
        suggestions.append("🟢 资产负债率 ({:.2f}%) 在合理范围内。".format(dar))

    npm = ratios.get("净利润率", 0)
    if npm > 20:
        suggestions.append("💰 净利润率 ({:.2f}%) 表现优秀，盈利能力强劲。".format(npm))
    elif npm > 10:
        suggestions.append("💰 净利润率 ({:.2f}%) 良好，盈利能力稳定。".format(npm))
    elif npm > 0:
        suggestions.append("📈 净利润率 ({:.2f}%) 偏低，建议分析成本结构寻找优化空间。".format(npm))

    roe = ratios.get("净资产收益率(ROE)", 0)
    if roe > 15:
        suggestions.append("🎯 ROE ({:.2f}%) 表现优秀，股东回报率高。".format(roe))
    elif roe > 0:
        suggestions.append("📉 ROE ({:.2f}%) 有提升空间，建议提高资产周转效率。".format(roe))

    er = ratios.get("费用占收入比", 0)
    if er > 80:
        suggestions.append("✂️ 费用占收入比高达 {:.2f}%，建议全面审查各项费用开支，实施成本控制。".format(er))
    elif er > 60:
        suggestions.append("📋 费用占收入比 {:.2f}%，关注费用增长趋势，控制不必要支出。".format(er))

    return suggestions


def get_trend_data(months: int = 6) -> Dict[str, List[Any]]:
    today = date.today()
    result: Dict[str, List[Any]] = {"labels": [], "revenue": [], "expense": [], "profit": []}

    for i in range(months - 1, -1, -1):
        m = today.month - i
        y = today.year
        while m <= 0:
            m += 12
            y -= 1

        start = date(y, m, 1)
        if m == 12:
            end = date(y + 1, 1, 1) - timedelta(days=1)
        else:
            end = date(y, m + 1, 1) - timedelta(days=1)

        label = f"{y}-{m:02d}"
        result["labels"].append(label)

        accounts = load_accounts_from_db()
        balances = calc_balances(accounts, as_of=end.isoformat())

        income_sum = sum(balances.get(a['code'], Decimal('0')) for a in accounts if a['category'] == 'income')
        expense_sum = sum(balances.get(a['code'], Decimal('0')) for a in accounts if a['category'] == 'expense')

        result["revenue"].append(float(income_sum))
        result["expense"].append(float(expense_sum))
        result["profit"].append(float(income_sum - expense_sum))

    return result


def predict_next_month(months: int = 3) -> List[Dict[str, Any]]:
    trend = get_trend_data(months)
    revenues = trend["revenue"]
    expenses = trend["expense"]

    if len(revenues) < 2:
        return []

    avg_revenue = sum(revenues) / len(revenues)
    avg_expense = sum(expenses) / len(expenses)

    if len(revenues) >= 3:
        rev_growth = (revenues[-1] - revenues[0]) / len(revenues)
        exp_growth = (expenses[-1] - expenses[0]) / len(expenses)
    else:
        rev_growth = 0
        exp_growth = 0

    predictions = []
    last_label = trend["labels"][-1] if trend["labels"] else f"{date.today().year}-{date.today().month:02d}"
    for i in range(1, months + 1):
        pred_rev = avg_revenue + rev_growth * i
        pred_exp = avg_expense + exp_growth * i
        pred_profit = pred_rev - pred_exp
        predictions.append({
            "period": f"{last_label}+{i}M",
            "predicted_revenue": round(pred_rev, 2),
            "predicted_expense": round(pred_exp, 2),
            "predicted_profit": round(pred_profit, 2),
        })

    return predictions


def get_template_names() -> list[str]:
    """Return list of available voucher template names."""
    from .constants import VOUCHER_TEMPLATES
    return list(VOUCHER_TEMPLATES.keys())


def apply_voucher_template(
    template_name: str, amount: float, date_str: Optional[str] = None
) -> Optional[str]:
    """Apply a voucher template to auto-create a voucher."""
    from .constants import VOUCHER_TEMPLATES

    if template_name not in VOUCHER_TEMPLATES:
        return None

    tpl = VOUCHER_TEMPLATES[template_name]
    if not date_str:
        date_str = datetime.now().strftime("%Y-%m-%d")

    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        return None

    year, month = dt.year, dt.month
    conn = get_conn()
    try:
        v_no = next_voucher_no(year, month)
        conn.execute(
            "INSERT INTO vouchers (voucher_no, date, summary, fiscal_year, fiscal_month) VALUES (?,?,?,?,?)",
            (v_no, date_str, tpl["summary"], year, month),
        )
        vid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        for code, direction in tpl["entries"]:
            debit = float(amount) if direction == "d" else 0
            credit = float(amount) if direction == "c" else 0
            conn.execute(
                "INSERT INTO journal_entries (voucher_id, account_code, debit, credit) VALUES (?,?,?,?)",
                (vid, code, debit, credit),
            )

        conn.commit()
        conn.close()
        return v_no
    except Exception:
        conn.rollback()
        conn.close()
        return None
